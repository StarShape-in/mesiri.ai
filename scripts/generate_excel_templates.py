import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ==========================================
# EXECUTIVE COLOR PALETTE (MERCON PREMIUM)
# ==========================================
DARK_NAVY = "0F172A"       # Slate 900 - Primary Brand Banner
DEEP_INDIGO = "1E1B4B"     # Indigo 950 - Sub Banners
HEADER_BG = "312E81"       # Indigo 900 - Table Headers
HEADER_REQ_BG = "4338CA"   # Indigo 600 - Required Field Header Accent
CARD_BG = "F1F5F9"         # Slate 100 - Top Summary Cards
CARD_BORDER = "CBD5E1"     # Slate 300 - Summary Card Border
ZEBRA_EVEN = "FFFFFF"      # White
ZEBRA_ODD = "F8FAFC"       # Slate 50 - Soft Alternating Tint
BORDER_SOFT = "E2E8F0"     # Slate 200 - Cell Border
REQUIRED_RED = "DC2626"    # Red accent for required badge text

# FONTS
font_brand = Font(name="Segoe UI", size=16, bold=True, color="FFFFFF")
font_sub_brand = Font(name="Segoe UI", size=10, italic=False, color="94A3B8")
font_card_title = Font(name="Segoe UI", size=9, bold=True, color="64748B")
font_card_value = Font(name="Segoe UI", size=12, bold=True, color="0F172A")

font_header_req = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
font_header_opt = Font(name="Segoe UI", size=10, bold=True, color="E2E8F0")

font_data = Font(name="Segoe UI", size=10, color="0F172A")
font_data_center = Font(name="Segoe UI", size=10, color="0F172A")
font_sample = Font(name="Segoe UI", size=10, italic=True, color="475569")

# FILLS
fill_brand = PatternFill(start_color=DARK_NAVY, end_color=DARK_NAVY, fill_type="solid")
fill_card = PatternFill(start_color=CARD_BG, end_color=CARD_BG, fill_type="solid")
fill_header_req = PatternFill(start_color=HEADER_REQ_BG, end_color=HEADER_REQ_BG, fill_type="solid")
fill_header_opt = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
fill_odd = PatternFill(start_color=ZEBRA_ODD, end_color=ZEBRA_ODD, fill_type="solid")
fill_even = PatternFill(start_color=ZEBRA_EVEN, end_color=ZEBRA_EVEN, fill_type="solid")

# BORDERS
border_cell = Border(
    left=Side(style='thin', color=BORDER_SOFT),
    right=Side(style='thin', color=BORDER_SOFT),
    top=Side(style='thin', color=BORDER_SOFT),
    bottom=Side(style='thin', color=BORDER_SOFT)
)

border_card = Border(
    left=Side(style='medium', color=CARD_BORDER),
    right=Side(style='medium', color=CARD_BORDER),
    top=Side(style='medium', color=CARD_BORDER),
    bottom=Side(style='medium', color=CARD_BORDER)
)

border_header = Border(
    left=Side(style='thin', color="475569"),
    right=Side(style='thin', color="475569"),
    top=Side(style='medium', color=DARK_NAVY),
    bottom=Side(style='medium', color=DARK_NAVY)
)

# ALIGNMENTS
align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
align_right = Alignment(horizontal='right', vertical='center', wrap_text=True)


def apply_executive_instructions(ws):
    ws.title = "📌 Guide & Overview"
    ws.views.sheetView[0].showGridLines = True

    # 1. Main Header Banner (Row 1-2)
    ws.merge_cells("A1:G2")
    banner = ws["A1"]
    banner.value = "🏢 MERCON LOGISTICS  |  ENTERPRISE FLEET & DRIVER IMPORT SPECIFICATION"
    banner.font = font_brand
    banner.fill = fill_brand
    banner.alignment = align_center

    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 24

    # 2. Subtitle Bar
    ws.merge_cells("A3:G3")
    sub = ws["A3"]
    sub.value = "Official onboarding worksheet for Company Owners & Fleet Administrators. Follow instructions below before returning the file."
    sub.font = font_sub_brand
    sub.alignment = align_center
    ws.row_dimensions[3].height = 20

    # 3. KPI / Status Cards (Row 5-6)
    cards = [
        ("A5:B6", "STATUS", "100% Import Ready", "A5"),
        ("C5:D6", "MANDATORY FIELDS", "Marked with Red Asterisk (*)", "C5"),
        ("E5:G6", "SUPPORT & INGESTION", "support@mercon-logistics.com", "E5")
    ]

    for cell_range, title, val, top_left in cards:
        ws.merge_cells(cell_range)
        top_cell = ws[top_left]
        top_cell.value = f"{title}\n{val}"
        top_cell.font = Font(name="Segoe UI", size=10, bold=True, color="1E293B")
        top_cell.fill = fill_card
        top_cell.alignment = align_center
        top_cell.border = border_card

    ws.row_dimensions[5].height = 20
    ws.row_dimensions[6].height = 20

    # 4. Instructions Content Table
    sections = [
        ("➊ STEP 1: REVIEW REQUIRED COLUMNS", [
            "• Columns with a red label and asterisk (*) are MANDATORY.",
            "• Required Driver fields: First Name, Last Name, Primary Phone, License Number, Expiry Date.",
            "• Required Vehicle fields: Plate Number, Asset Type, Capacity (KG)."
        ]),
        ("➋ STEP 2: USE INTEGRATED DROPDOWNS", [
            "• Fields such as Asset Type (Flatbed, Reefer, Box, Tanker) have built-in Excel picklists.",
            "• Please select directly from the dropdown to prevent spelling mismatches."
        ]),
        ("➌ STEP 3: ICCES GPS TRACKING INTEGRATION", [
            "• For every truck equipped with an ICCES telematics unit, enter its unique ICCES Device ID.",
            "• MERCON will automatically start pulling live GPS position, speed, and odometer upon import."
        ]),
        ("➍ STEP 4: SUBMIT FOR AUTOMATED PARSING", [
            "• Save this workbook in standard Excel (.xlsx) or (.csv) format.",
            "• Upload directly in the MERCON Web Dashboard under 'Bulk Fleet Import'."
        ])
    ]

    current_row = 9
    for sec_title, points in sections:
        ws.cell(row=current_row, column=1, value=sec_title).font = Font(name="Segoe UI", size=11, bold=True, color="3730A3")
        current_row += 1
        for pt in points:
            ws.cell(row=current_row, column=2, value=pt).font = Font(name="Segoe UI", size=10, color="334155")
            current_row += 1
        current_row += 1

    ws.column_dimensions['A'].width = 38
    ws.column_dimensions['B'].width = 85


def apply_drivers_sheet(ws):
    ws.title = "👨‍✈️ Drivers Directory"
    ws.views.sheetView[0].showGridLines = True

    # 1. Header Banner
    ws.merge_cells("A1:G2")
    banner = ws["A1"]
    banner.value = "👨‍✈️ DRIVER PERSONNEL DIRECTORY  —  BULK DATA ENTRY"
    banner.font = font_brand
    banner.fill = fill_brand
    banner.alignment = align_center

    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 24

    # 2. Executive KPI Summary Cards (Row 4-5)
    ws.merge_cells("A4:B5")
    c1 = ws["A4"]
    c1.value = "MODULE\nDriver Master Directory"
    c1.font = font_card_title
    c1.fill = fill_card
    c1.alignment = align_center
    c1.border = border_card

    ws.merge_cells("C4:E5")
    c2 = ws["C4"]
    c2.value = "PRIMARY KEY & VEHICLE ASSIGNMENT\nPhone Number, License & Assigned Vehicle Plate"
    c2.font = font_card_title
    c2.fill = fill_card
    c2.alignment = align_center
    c2.border = border_card

    ws.merge_cells("F4:G5")
    c3 = ws["F4"]
    c3.value = "INITIAL STATUS\nAuto-Defaulted to 'Available'"
    c3.font = font_card_title
    c3.fill = fill_card
    c3.alignment = align_center
    c3.border = border_card

    ws.row_dimensions[4].height = 18
    ws.row_dimensions[5].height = 18

    # 3. Table Column Headers (Row 7)
    headers = [
        ("Driver Ref / ID", "Optional ID (e.g. DRV-101)", False, 18),
        ("First Name *", "Mandatory First Name", True, 22),
        ("Last Name *", "Mandatory Last Name", True, 22),
        ("Primary Phone *", "Unique Mobile Number", True, 26),
        ("License Number *", "Official Driving License", True, 24),
        ("License Expiry *", "YYYY-MM-DD", True, 20),
        ("Assigned Vehicle Plate", "License Plate of assigned truck (e.g. 8492-RKA)", False, 28)
    ]

    ws.row_dimensions[7].height = 30

    for col_idx, (header_text, desc, is_req, col_width) in enumerate(headers, start=1):
        cell = ws.cell(row=7, column=col_idx, value=header_text)
        cell.font = font_header_req if is_req else font_header_opt
        cell.fill = fill_header_req if is_req else fill_header_opt
        cell.alignment = align_center
        cell.border = border_header
        
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = col_width

    # 4. Sample Pre-filled Rows (Rows 8 to 10)
    samples = [
        ["DRV-101", "Ahmed", "Al-Mansoor", "+966 50 123 4567", "DL-98765432", "2028-12-31", "8492-RKA"],
        ["DRV-102", "Tariq", "Hassan", "+966 50 987 6543", "DL-12345678", "2027-06-15", "3104-LMN"],
        ["DRV-103", "John", "Doe", "+1 555 019 2834", "DL-55443322", "2029-09-20", "9912-XYZ"]
    ]

    for row_idx, sample_data in enumerate(samples, start=8):
        ws.row_dimensions[row_idx].height = 24
        for col_idx, val in enumerate(sample_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = font_sample
            cell.alignment = align_center if col_idx in [1, 6, 7] else align_left
            cell.border = border_cell
            cell.fill = fill_odd if row_idx % 2 == 1 else fill_even

    # 5. Empty Input Rows (Rows 11 to 60)
    for row_idx in range(11, 61):
        ws.row_dimensions[row_idx].height = 22
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = font_data
            cell.border = border_cell
            cell.fill = fill_odd if row_idx % 2 == 1 else fill_even
            if col_idx in [1, 6, 7]:
                cell.alignment = align_center
                if col_idx == 6:
                    cell.number_format = 'YYYY-MM-DD'
            elif col_idx == 4:
                cell.number_format = '@'  # Text format for phone
                cell.alignment = align_left
            else:
                cell.alignment = align_left


def apply_vehicles_sheet(ws):
    ws.title = "🚛 Vehicles & Trailers Directory"
    ws.views.sheetView[0].showGridLines = True

    # 1. Header Banner
    ws.merge_cells("A1:I2")
    banner = ws["A1"]
    banner.value = "🚛 VEHICLE & TRAILER FLEET DIRECTORY  —  BULK DATA ENTRY"
    banner.font = font_brand
    banner.fill = fill_brand
    banner.alignment = align_center

    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 24

    # 2. Executive KPI Summary Cards (Row 4-5)
    ws.merge_cells("A4:B5")
    c1 = ws["A4"]
    c1.value = "MODULE\nFleet & Telematics Directory"
    c1.font = font_card_title
    c1.fill = fill_card
    c1.alignment = align_center
    c1.border = border_card

    ws.merge_cells("C4:F5")
    c2 = ws["C4"]
    c2.value = "TELEMATICS & DRIVER ASSIGNMENT\nEnter ICCES Device ID & Assigned Driver Phone / Name"
    c2.font = font_card_title
    c2.fill = fill_card
    c2.alignment = align_center
    c2.border = border_card

    ws.merge_cells("G4:I5")
    c3 = ws["G4"]
    c3.value = "INITIAL STATUS\nAuto-Defaulted to 'Available'"
    c3.font = font_card_title
    c3.fill = fill_card
    c3.alignment = align_center
    c3.border = border_card

    ws.row_dimensions[4].height = 18
    ws.row_dimensions[5].height = 18

    # 3. Table Column Headers (Row 7)
    headers = [
        ("Vehicle Ref / ID", "Optional ID (e.g. TRK-101)", False, 18),
        ("Plate Number *", "Mandatory License Plate", True, 22),
        ("Asset Type *", "Flatbed / Reefer / Box / Tanker", True, 22),
        ("Capacity (KG) *", "Payload Capacity in KG", True, 20),
        ("Current Odometer (KM)", "Current Odometer", False, 22),
        ("ICCES Device ID", "ICCES GPS Device Serial ID", False, 25),
        ("Assigned Driver Phone / Name", "Phone or Name of assigned driver", False, 28),
        ("Trailer Number", "Linked Trailer Number", False, 20),
        ("Trailer Capacity (KG)", "Trailer Payload Capacity in KG", False, 24)
    ]

    ws.row_dimensions[7].height = 30

    for col_idx, (header_text, desc, is_req, col_width) in enumerate(headers, start=1):
        cell = ws.cell(row=7, column=col_idx, value=header_text)
        cell.font = font_header_req if is_req else font_header_opt
        cell.fill = fill_header_req if is_req else fill_header_opt
        cell.alignment = align_center
        cell.border = border_header
        
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = col_width

    # 4. Sample Pre-filled Rows (Rows 8 to 10)
    samples = [
        ["TRK-101", "8492-RKA", "Flatbed", 25000, 45000, "06670881", "+966 50 123 4567 (Ahmed)", "TRL-402", 30000],
        ["TRK-102", "3104-LMN", "Reefer", 20000, 32000, "351777090213198", "+966 50 987 6543 (Tariq)", "TRL-405", 22000],
        ["TRK-103", "9912-XYZ", "Tanker", 35000, 88000, "08085338", "+1 555 019 2834 (John)", "TRL-409", 35000]
    ]

    for row_idx, sample_data in enumerate(samples, start=8):
        ws.row_dimensions[row_idx].height = 24
        for col_idx, val in enumerate(sample_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = font_sample
            cell.border = border_cell
            cell.fill = fill_odd if row_idx % 2 == 1 else fill_even
            
            if col_idx in [4, 5, 9]:
                cell.alignment = align_right
                cell.number_format = '#,##0'
            elif col_idx in [1, 3, 6, 8]:
                cell.alignment = align_center
            else:
                cell.alignment = align_left

    # 5. Empty Input Rows (Rows 11 to 60)
    for row_idx in range(11, 61):
        ws.row_dimensions[row_idx].height = 22
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = font_data
            cell.border = border_cell
            cell.fill = fill_odd if row_idx % 2 == 1 else fill_even
            
            if col_idx in [4, 5, 9]:
                cell.alignment = align_right
                cell.number_format = '#,##0'
            elif col_idx in [1, 3, 6, 8]:
                cell.alignment = align_center
            else:
                cell.alignment = align_left

    # Data Validation for Asset Type
    dv_asset = DataValidation(type="list", formula1='"Flatbed,Reefer,Box,Tanker"', allow_blank=True)
    dv_asset.error = 'Please pick a valid Asset Type from the list'
    dv_asset.errorTitle = 'Invalid Asset Type'
    ws.add_data_validation(dv_asset)
    dv_asset.add("C8:C60")


def generate_all_professional_templates():
    output_dirs = [
        r"d:\Mercon\docs\templates",
        r"d:\Mercon\frontend\public\templates"
    ]

    # 1. Dedicated Drivers Import Workbook
    wb_drivers = openpyxl.Workbook()
    ws_d_inst = wb_drivers.active
    apply_executive_instructions(ws_d_inst)
    ws_d_sheet = wb_drivers.create_sheet()
    apply_drivers_sheet(ws_d_sheet)

    # 2. Dedicated Vehicles Import Workbook
    wb_vehicles = openpyxl.Workbook()
    ws_v_inst = wb_vehicles.active
    apply_executive_instructions(ws_v_inst)
    ws_v_sheet = wb_vehicles.create_sheet()
    apply_vehicles_sheet(ws_v_sheet)

    for out_dir in output_dirs:
        os.makedirs(out_dir, exist_ok=True)
        
        driver_path = os.path.join(out_dir, "MERCON_Drivers_Import_Template.xlsx")
        vehicle_path = os.path.join(out_dir, "MERCON_Vehicles_Import_Template.xlsx")

        wb_drivers.save(driver_path)
        wb_vehicles.save(vehicle_path)

        print(f"[SUCCESS] Generated 2 requested Excel templates in: {out_dir}")

if __name__ == "__main__":
    generate_all_professional_templates()
